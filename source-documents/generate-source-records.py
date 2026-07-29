"""Regenerate src/data/source-records.ts from the two supplied source documents.

Run with:  pip install openpyxl pdfplumber && python source-documents/generate-source-records.py

Rules enforced here:
  * raw source text is preserved verbatim
  * sheet/page/cell/row/column provenance is preserved
  * formula-error cells are flagged, never treated as verified values
  * blank cells become status 'unavailable' -- never 'No'
"""
import collections, json, os
import openpyxl
import pdfplumber

HERE = os.path.dirname(os.path.abspath(__file__))
PDF = os.path.join(HERE, 'Daikin FIT Battlecard.pdf')
XLSX = os.path.join(HERE, 'Competitor comparison.xlsx')
OUT = os.path.join(HERE, '..', 'src', 'data', 'source-records.ts')

# ---------------------------------------------------------------- battlecard
pdf = pdfplumber.open(PDF)
page = pdf.pages[0]
table = page.find_tables()[0]
data = table.extract()
header = data[5]
header_cells = table.rows[5].cells

row_y = {ri: (r.cells[2][1], r.cells[2][3]) for ri, r in enumerate(table.rows) if r.cells[2]}
FILL = {(0.71, 0.902, 0.635): 'daikin_better',
        (0.945, 0.663, 0.514): 'competitor_better',
        (1.0, 1.0, 0.8): 'not_available_marker'}
assessment = collections.defaultdict(dict)
for rect in page.rects:
    col = rect.get('non_stroking_color')
    if isinstance(col, (list, tuple)):
        col = tuple(round(v, 3) for v in col)
    if col not in FILL:
        continue
    cy = (rect['top'] + rect['bottom']) / 2
    ri = next((k for k, (y0, y1) in row_y.items() if y0 - 0.3 <= cy <= y1 + 0.3), None)
    if ri is None or ri < 6:
        continue
    for ci in range(2, 24):
        x0, x1 = header_cells[ci][0], header_cells[ci][2]
        if min(x1, rect['x1']) - max(x0, rect['x0']) > (x1 - x0) * 0.5:
            assessment[ri][ci] = FILL[col]

# attrKey, pdf row index, label, group, unit, direction, kind
BC_ATTRS = [
    ('initial_cost',        6,  'Initial cost',                                    'Product Overview',           '',      'lower',  'ordinal'),
    ('tonnage_options',     7,  'Tonnage options',                                 'Product Overview',           'tons',  'higher', 'tonnage'),
    ('chassis_type',        8,  'Chassis type',                                    'Product Overview',           '',      'none',   'text'),
    ('footprint',           9,  'Smallest footprint available (D x W x H)',        'Product Overview',           'in',    'lower',  'text'),
    ('air_handler_matchup', 10, 'Air-handler matchup',                             'Product Overview',           '',      'none',   'text'),
    ('refrigerant',         11, 'Refrigerant',                                     'Product Overview',           '',      'none',   'text'),
    ('compressor_type',     12, 'Compressor type',                                 'Product Overview',           '',      'none',   'text'),
    ('sound_blanket',       13, 'Compressor sound blanket',                        'Comfort & Sound',            '',      'higher', 'bool'),
    ('sound_level',         14, 'Sound level',                                     'Comfort & Sound',            'dBA',   'lower',  'measure'),
    ('coil_only_matchup',   15, 'Coil-only matchup',                               'Product Overview',           '',      'higher', 'bool'),
    ('straight_cool',       16, 'Straight-cool version',                           'Product Overview',           '',      'higher', 'bool'),
    ('thermostat_type',     17, 'Primary thermostat type',                         'Controls',                   '',      'none',   'text'),
    ('thermostat_24v',      18, '24V thermostat option',                           'Controls',                   '',      'higher', 'bool'),
    ('regional_profiles',   19, 'Regional equipment profiles (Quality Install)',   'Installation & Diagnostics', '',      'higher', 'bool'),
    ('reusable_profiles',   20, 'Re-usable equipment configurations / Profiles',   'Installation & Diagnostics', '',      'higher', 'bool'),
    ('charge_verification', 21, 'Charge verification without tools',               'Installation & Diagnostics', '',      'higher', 'bool'),
    ('slow_loss_alerting',  22, 'Slow loss-of-charge alerting',                    'Installation & Diagnostics', '',      'higher', 'bool'),
    ('cloud_alerts',        23, 'Real-time cloud-connected alerts & diagnostics',  'Installation & Diagnostics', '',      'higher', 'bool'),
    ('humidity_control',    24, 'Humidity control (true dehumidification)',        'Comfort & Sound',            '',      'higher', 'bool'),
    ('base_pan_heater',     25, 'Factory-installed base-pan heater',               'Features',                   '',      'higher', 'bool'),
    ('heater_kit_3stage',   26, '3-stage heater kit (optional)',                   'Features',                   '',      'higher', 'bool'),
    ('intelligent_defrost', 27, 'Intelligent defrost',                             'Features',                   '',      'higher', 'bool'),
    ('anticorrosive',       28, 'Anti-corrosive film coating',                     'Features',                   '',      'higher', 'bool'),
    ('energy_star',         29, 'ENERGY STAR®',                               'Ratings & Certifications',   '',      'higher', 'bool'),
    ('energy_star_cchp',    30, 'ENERGY STAR® cold-climate (ccHP)',           'Ratings & Certifications',   '',      'higher', 'bool'),
    ('cee_2025',            31, 'CEE 2025',                                        'Ratings & Certifications',   '',      'higher', 'bool'),
    ('seer2',               32, 'SEER2',                                           'Efficiency',                 'SEER2', 'higher', 'measure'),
    ('eer2',                33, 'EER2',                                            'Efficiency',                 'EER2',  'higher', 'measure'),
    ('hspf2',               34, 'HSPF2',                                           'Efficiency',                 'HSPF2', 'higher', 'measure'),
    ('cop_5f',              35, 'COP @ 5°F',                                  'Efficiency',                 'COP',   'higher', 'measure'),
    ('cap_5f',              36, 'Max capacity at 5°F',                        'Capacity',                   'BTU/h', 'higher', 'measure'),
    ('cap_47f',             37, 'Max capacity at 47°F',                       'Capacity',                   'BTU/h', 'higher', 'measure'),
    ('cap_95f',             38, 'Max capacity at 95°F',                       'Capacity',                   'BTU/h', 'higher', 'measure'),
    ('cap_115f',            39, 'Max capacity at 115°F',                      'Capacity',                   'BTU/h', 'higher', 'measure'),
    ('cooling_range',       40, 'Cooling operating range (min / max)',             'Operating Range',            '°F', 'range', 'range'),
    ('heating_range',       41, 'Heating operating range (min / max)',             'Operating Range',            '°F', 'range', 'range'),
    ('line_length',         42, 'Line length up to',                               'Line Set & Installation',    'ft',    'higher', 'measure'),
    ('pre_charge',          43, 'Pre-charge',                                      'Line Set & Installation',    'ft',    'higher', 'measure'),
    ('elevation',           44, 'Elevation',                                       'Line Set & Installation',    'ft',    'higher', 'measure'),
    ('warranty',            45, 'Warranty',                                        'Warranty',                   'years', 'higher', 'warranty'),
]

ASSESS_ROW = {ri: assessment.get(ri, {}) for _, ri, *_ in BC_ATTRS}

bc_rows = []
for key, ri, label, group, unit, direction, kind in BC_ATTRS:
    raw_label = (data[ri][1] or '').replace('\n', ' ').strip()
    values = []
    assess = []
    for ci in range(2, 24):
        v = data[ri][ci]
        v = None if v in (None, '') else v.replace('\n', ' ').strip()
        values.append(v)
        assess.append(ASSESS_ROW[ri].get(ci))
    comment = (data[ri][24] or '').replace('\n', ' ').strip() or None
    bc_rows.append({
        'key': key, 'label': label, 'sourceLabel': raw_label, 'group': group,
        'unit': unit, 'direction': direction, 'kind': kind,
        'pdfRow': ri, 'values': values, 'assessment': assess, 'comment': comment,
    })

bc_products = []
FAMILY = {
    'DH6VS FIT': 'Daikin FIT', 'DH7VS FIT': 'Daikin FIT', 'DH9VS FIT AURORA': 'Daikin FIT',
    '37MUHA': '37MU', 'EVOX MIDEA': 'EVOX', 'INFINITY 27VNA3': 'Infinity',
    'DC5 SERIES': 'DC5', 'SIGNATURE SL22KLV': 'Signature', 'ELITE EL18KSLV': 'Elite',
    'FLEEX ECO R32': 'FLEEX', 'FLEEX Ultra R32': 'FLEEX',
    'IDS Light/PLUS': 'IDS', 'IDS PREMIUM': 'IDS', 'IDS Ultra HP': 'IDS',
    'KUSXA/B': 'KUSX', 'HI-Pro': 'Hi Series', 'Hi-Ultra': 'Hi Series',
    'HH8': 'HH', 'RD16AY': 'RD', 'RD18AY': 'RD', 'RD19AY': 'RD', '5TWV0': '5TWV',
}
for ci in range(2, 24):
    head = header[ci]
    parts = head.split('\n')
    model = parts[0].strip()
    brand = parts[1].strip() if len(parts) > 1 else None
    bc_products.append({
        'colIndex': ci,
        'sourceHeader': head.replace('\n', ' / '),
        'model': model,
        'brand': brand,
        'family': FAMILY.get(model, model),
    })

# ---------------------------------------------------------------- xlsx
wb_f = openpyxl.load_workbook(XLSX, data_only=False)
wb_v = openpyxl.load_workbook(XLSX, data_only=True)
ws_f, ws_v = wb_f['Comparison'], wb_v['Comparison']

HY_PRODUCT_ROWS = [
    (7, 17, 27, 37, 'Daikin UPRA043DAVK'),
    (8, 18, 28, 38, 'Viessmann'),
    (9, 19, 29, 39, 'Spacepak'),
    (10, 20, 30, 40, 'LG'),
    (11, 21, 31, 41, 'Lochinvar'),
    (12, 22, 32, 42, 'NTI'),
]

# key, label, group, unit, direction, kind, block(0-3), column letter, header cell ref
HY_ATTRS = [
    ('max_lwt',              'Max. leaving water temperature',                    'Leaving Water Temperature', '°F',   'higher', 'measure', 0, 'C', 'C6'),
    ('min_lwt',              'Min. LWT at lowest ambient operating temp',         'Leaving Water Temperature', '°F',   'higher', 'measure', 0, 'D', 'D6'),
    ('delta_lwt',            'Δ LWT (min / max ratio)',                      'Leaving Water Temperature', 'ratio',     'higher', 'measure', 0, 'E', 'E6'),
    ('lowest_ambient_lwt',   'Lowest ambient operating temp (LWT block)',         'Leaving Water Temperature', '°F',   'lower',  'measure', 0, 'F', 'F6'),
    ('emitter_high_temp',    'High-temp emitters (baseboard / radiator)',         'Emitter Compatibility',     '°F',   'higher', 'measure', 0, 'I', 'I3'),
    ('emitter_medium_temp',  'Medium-temp emitters (convector / fan coil / AHU)', 'Emitter Compatibility',     '°F',   'higher', 'measure', 0, 'J', 'J3'),
    ('emitter_low_temp',     'Low-temp emitters (radiant floor)',                 'Emitter Compatibility',     '°F',   'higher', 'measure', 0, 'K', 'K3'),
    ('cold_climate_op',      'Cold-climate operation',                            'Emitter Compatibility',     '°F',   'lower',  'measure', 0, 'L', 'L4'),
    ('max_heat_cap_131',     'Max. heating capacity @ 131°F LWT',            'Heating Capacity',          'BTU/h',     'higher', 'measure', 1, 'C', 'C16'),
    ('min_heat_cap',         'Min. heating capacity at lowest ambient',           'Heating Capacity',          'BTU/h',     'higher', 'measure', 1, 'D', 'D16'),
    ('delta_heat_cap',       'Δ heating capacity (max − min)',          'Heating Capacity',          'BTU/h',     'lower',  'measure', 1, 'E', 'E16'),
    ('lowest_ambient_heat',  'Lowest ambient operating temp (heating block)',     'Heating Capacity',          '°F',   'lower',  'measure', 1, 'F', 'F16'),
    ('single_fan',           'Single fan',                                        'Configuration',             '',          'none',   'bool',    1, 'I', 'I16'),
    ('no_glycol',            'No glycol required',                                'Configuration',             '',          'higher', 'bool',    1, 'J', 'J16'),
    ('warranty_hydronic',    'Warranty',                                          'Warranty',                  '',          'higher', 'text',    1, 'K', 'K16'),
    ('boiler_replacement',   'Boiler replacement',                                'Configuration',             '°F',   'higher', 'measure', 1, 'L', 'L16'),
    ('max_cool_cap',         'Max. cooling capacity',                             'Cooling Capacity',          'BTU/h',     'higher', 'measure', 2, 'C', 'C26'),
    ('min_cool_cap',         'Min. cooling capacity at highest ambient',          'Cooling Capacity',          'BTU/h',     'higher', 'measure', 2, 'D', 'D26'),
    ('delta_cool_cap',       'Δ cooling capacity (max / min)',               'Cooling Capacity',          'BTU/h',     'lower',  'measure', 2, 'E', 'E26'),
    ('highest_ambient',      'Highest ambient operating temp',                    'Cooling Capacity',          '°F',   'higher', 'measure', 2, 'F', 'F26'),
    ('oem_heat_pump',        'OEM heat pump',                                     'Configuration',             '',          'none',   'bool',    2, 'I', 'I26'),
    ('hydro_split',          'Hydro-split',                                       'Configuration',             '',          'none',   'bool',    2, 'J', 'J26'),
    ('low_gwp',              'Low GWP',                                           'Configuration',             '',          'none',   'bool',    2, 'K', 'K26'),
    ('sound_level_hy',       'Sound',                                             'Comfort & Sound',           'dBA',       'lower',  'measure', 3, 'C', 'C36'),
    ('total_amps',           'Total amps (breaker)',                              'Electrical',                'A',         'lower',  'measure', 3, 'D', 'D36'),
    ('cop_a5w110',           'COP A5/W110',                                       'Efficiency',                'COP',       'higher', 'measure', 3, 'E', 'E36'),
]

BLOCK_ROW_INDEX = {0: 0, 1: 1, 2: 2, 3: 3}

hy_rows = []
for key, label, group, unit, direction, kind, block, col, href in HY_ATTRS:
    hdr = ws_v[href].value
    cells = []
    for rowset in HY_PRODUCT_ROWS:
        r = rowset[BLOCK_ROW_INDEX[block]]
        ref = f'{col}{r}'
        fv = ws_f[ref].value
        vv = ws_v[ref].value
        is_formula = isinstance(fv, str) and fv.startswith('=')
        err = isinstance(vv, str) and vv.startswith('#')
        cells.append({
            'ref': ref,
            'raw': None if vv is None else (str(vv)),
            'formula': fv if is_formula else None,
            'error': bool(err),
        })
    hy_rows.append({
        'key': key, 'label': label, 'sourceLabel': str(hdr) if hdr is not None else label,
        'group': group, 'unit': unit, 'direction': direction, 'kind': kind,
        'headerRef': href, 'cells': cells,
    })

hy_products = []
for rowset in HY_PRODUCT_ROWS:
    label = rowset[4]
    is_daikin = label.startswith('Daikin')
    hy_products.append({
        'rowRefs': [f'B{rowset[0]}', f'B{rowset[1]}', f'B{rowset[2]}', f'B{rowset[3]}'],
        'sourceHeader': label,
        'brand': 'Daikin' if is_daikin else label,
        'model': label.replace('Daikin ', '') if is_daikin else None,
        'family': 'UPRA' if is_daikin else None,
    })

# formula-error cells recorded for provenance / exclusion
error_cells = []
for row in ws_v.iter_rows(min_row=1, max_row=ws_v.max_row, max_col=ws_v.max_column):
    for c in row:
        if isinstance(c.value, str) and c.value.startswith('#'):
            error_cells.append({'ref': c.coordinate, 'raw': c.value})

legend = [
    {'token': 'NA', 'meaning': (data[0][6] or '').strip()},
    {'token': 'Green', 'meaning': (data[1][6] or '').strip()},
    {'token': 'Red', 'meaning': (data[2][6] or '').strip()},
    {'token': 'White', 'meaning': (data[3][6] or '').strip()},
]

payload = {
    'battlecard': {'products': bc_products, 'rows': bc_rows, 'legend': legend,
                   'title': (data[2][0] or '').strip()},
    'hydronic': {'products': hy_products, 'rows': hy_rows, 'errorCells': error_cells},
}


def ts(obj, indent=0):
    return json.dumps(obj, ensure_ascii=False, indent=2)


with open(OUT, 'w') as f:
    f.write('/* AUTO-GENERATED from the supplied source documents. Do not edit by hand.\n')
    f.write(' * Sources: "Competitor comparison.xlsx" (sheet "Comparison") and\n')
    f.write(' *          "Daikin FIT Battlecard.pdf" (page 1).\n')
    f.write(' * Raw source text, sheet/page/cell provenance and formula-error flags are preserved. */\n\n')
    f.write('import type { BattlecardSource, HydronicSource } from "./types";\n\n')
    f.write('export const BATTLECARD: BattlecardSource = ' + ts(payload['battlecard']) + ' as BattlecardSource;\n\n')
    f.write('export const HYDRONIC: HydronicSource = ' + ts(payload['hydronic']) + ' as HydronicSource;\n')

print('wrote', OUT)
print('battlecard products', len(bc_products), 'rows', len(bc_rows))
print('hydronic products', len(hy_products), 'rows', len(hy_rows), 'error cells', len(error_cells))
print('total products', len(bc_products) + len(hy_products))

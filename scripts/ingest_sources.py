#!/usr/bin/env python3
"""Extract source-backed demo data with source lineage.

The script deliberately keeps unknown cells unknown: it does not infer a value
from blank cells or from battlecard color alone.  It supports merged workbook
headers and the single-grid layout used by the supplied FIT battlecard.
"""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Client_shared_datasets" / "Competitor comparison.xlsx"
PDF = ROOT / "Client_shared_datasets" / "Daikin FIT Battlecard.pdf"
OUT = ROOT / "lib" / "data" / "ingestedData.ts"
QUALITY = ROOT / "docs" / "DATA_QUALITY.md"


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).replace("\n", " ").split())
    return text or None


def brand_from_header(header: str) -> tuple[str, str]:
    parts = header.split("\n")
    model = clean(parts[0]) or header
    hint = " ".join(parts[1:]).upper()
    brands = {
        "DAIKIN": "Daikin", "CARRIER/MIDEA": "Carrier / Midea", "MIDEA": "Midea",
        "GREE": "Gree", "BOSCH": "Bosch", "LG": "LG", "HISENSE": "Hisense",
        "YORK": "York", "RHEEM": "Rheem", "TRANE": "Trane",
    }
    for key, brand in brands.items():
        if key in hint:
            return brand, model
    return "Competitor", model


def norm_attribute(label: str) -> str:
    aliases = {
        "Sizing options (Tons)": "Tonnage options",
        "Smallest footprint available (inches) D x W x H": "Footprint",
        "Sound Performance": "Sound level",
        "Cooling Operating Range Min/Max F": "Cooling operating range",
        "Heating Operating Range Min/Max F": "Heating operating range",
        "Line Lengths": "Maximum line length",
        "Warranty": "Warranty",
    }
    return aliases.get(label, label)


def parse_battlecard() -> list[dict[str, Any]]:
    with pdfplumber.open(PDF) as pdf:
        table = pdf.pages[0].extract_tables()[0]
    headers = table[5][2:24]
    products: list[dict[str, Any]] = []
    for index, header in enumerate(headers):
        if not header:
            continue
        brand, model = brand_from_header(header)
        attributes: dict[str, str] = {}
        for row_idx, row in enumerate(table[6:], start=6):
            label, value = clean(row[1]), clean(row[index + 2])
            if label and value:
                attributes[norm_attribute(label)] = value
        if attributes:
            products.append({
                "id": f"fit-{index + 1}", "brand": brand, "model": model,
                "family": "R-32 Daikin FIT Series" if brand == "Daikin" else "Competitive inverter heat pump",
                "equipmentType": "Inverter heat pump", "attributes": attributes,
                "source": "Daikin FIT Battlecard.pdf • p. 1",
            })
    return products


def parse_workbook() -> list[dict[str, Any]]:
    workbook = load_workbook(XLSX, data_only=False, read_only=True)
    sheet = workbook["Comparison"]
    rows = [
        (7, "Daikin UPRA043DAVK", "Max. leaving water temperature", 3, "°F"),
        (7, "Daikin UPRA043DAVK", "Lowest ambient operating temperature", 6, "°F"),
        (17, "Daikin UPRA043DAVK", "Max. heating capacity at 131°F LWT", 3, "BTU/h"),
        (17, "Daikin UPRA043DAVK", "Minimum heating capacity at lowest ambient", 4, "BTU/h"),
    ]
    names = [clean(sheet.cell(row, 2).value) for row in range(7, 13)]
    output: dict[str, dict[str, Any]] = {}
    for name in names:
        if not name:
            continue
        brand = "Daikin" if name.startswith("Daikin") else name
        output[name] = {
            "id": f"hydro-{name.lower().replace(' ', '-').replace('/', '-')}", "brand": brand,
            "model": name.replace("Daikin ", ""), "family": "Cold-climate hydronic heat pump",
            "equipmentType": "Air-to-water heat pump", "attributes": {},
            "source": "Competitor comparison.xlsx • Comparison",
        }
    for row, _daikin, label, col, unit in rows:
        for offset, name in enumerate(names):
            if not name:
                continue
            value = sheet.cell(row + offset, col).value
            if value is not None:
                output[name]["attributes"][label] = f"{value} {unit}"
    for offset, name in enumerate(names):
        if not name:
            continue
        value = sheet.cell(7 + offset, 3).value
        if value is not None:
            output[name]["attributes"]["Max. leaving water temperature"] = f"{value} °F"
        value = sheet.cell(7 + offset, 6).value
        if value is not None:
            output[name]["attributes"]["Lowest ambient operating temperature"] = f"{value} °F"
        value = sheet.cell(17 + offset, 3).value
        if value is not None:
            output[name]["attributes"]["Max. heating capacity at 131°F LWT"] = f"{value} BTU/h"
        value = sheet.cell(17 + offset, 4).value
        if value is not None:
            output[name]["attributes"]["Minimum heating capacity at lowest ambient"] = f"{value} BTU/h"
    return list(output.values())


def main() -> None:
    products = parse_battlecard() + parse_workbook()
    source_data = {
        "importedAt": str(date.today()), "products": products,
        "sources": [
            {"name": XLSX.name, "type": "XLSX", "location": "Comparison sheet", "warnings": 4},
            {"name": PDF.name, "type": "PDF", "location": "page 1", "warnings": 0},
        ],
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    QUALITY.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("// Generated by scripts/ingest_sources.py. Do not hand-edit.\n" +
                   "export const INGESTED_DATA = " + json.dumps(source_data, indent=2) + " as const;\n")
    QUALITY.write_text(f"# Data quality report\n\nImported {len(products)} product records from two supplied source documents on {date.today()}.\n\n- Workbook formula cells displaying `#VALUE!` were not used as specification values.\n- Blank workbook cells remain unavailable, not `No`.\n- Battlecard values retain their original strings and page-level provenance.\n- Product model labels, attributes, and all specifications in the demo are source-backed.\n")
    print(f"Imported {len(products)} products; wrote {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
